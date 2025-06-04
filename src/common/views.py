import openpyxl
from django.db.models import Q, QuerySet
from django.http import HttpResponse
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from requests import Response
from rest_framework import generics, status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.status import HTTP_200_OK

from api.constants import SortOderEnum
from utilities.permissions.custom_permissions import IsAuthenticated


class SortAndFilterViewSet(viewsets.ModelViewSet):
    model = None

    def get_queryset_by_filter(self, queryset):
        request = self.request
        user = request.user
        is_view_all = getattr(request, "is_view_all", None)
        if not is_view_all:
            queryset = queryset.filter(created_by=user)

        return self._generate_query_with_filter(
            queryset=queryset,
            search_fields=(
                self.model.SEARCH_FIELDS
                if self.model and getattr(self.model, "SEARCH_FIELDS", None)
                else {}
            ),
            search_fields_contains=(
                self.model.SEARCH_FIELDS_CONTAINS
                if self.model and getattr(self.model, "SEARCH_FIELDS_CONTAINS", None)
                else {}
            ),
            search_fields_custom=(
                self.model.SEARCH_FIELDS_CUSTOM
                if self.model and getattr(self.model, "SEARCH_FIELDS_CUSTOM", None)
                else {}
            ),
            query_params=request.query_params,
        )

    def get_queryset_by_sort(self, queryset):
        params = self.request.query_params
        sort_order = params.get("sort_order", None)
        sort_by = params.get("sort_by", None)
        return self._generate_query_with_sort(
            queryset=queryset,
            sort_order=sort_order,
            sort_by=sort_by,
        )

    @staticmethod
    def _generate_query_with_filter(
        queryset: QuerySet,
        search_fields: dict,
        search_fields_contains: dict,
        search_fields_custom: dict,
        query_params,
    ) -> QuerySet:
        filters = {}

        for key, value in search_fields.items():
            if key in query_params:
                filters[value] = query_params[key]

        for key, value in search_fields_contains.items():
            if key in query_params:
                filters[f"{value}__icontains"] = query_params[key]

        for key, value in search_fields_custom.items():
            if key in query_params:
                filters[value] = query_params[key]

        return queryset.filter(Q(**filters))

    @staticmethod
    def _generate_query_with_sort(
        queryset: QuerySet,
        sort_order: str,
        sort_by: str,
    ) -> QuerySet:
        if sort_order == SortOderEnum.ASC.value:
            queryset = queryset.order_by(f"{sort_by}")
        elif sort_order == SortOderEnum.DESC.value:
            queryset = queryset.order_by(f"-{sort_by}")
        return queryset


class ListAPI(generics.ListAPIView):
    serializer_class = None
    permission_classes = [IsAuthenticated]

    @property
    def model(self):
        return self.serializer_class.Meta.model

    def get(self, request, *args, **kwargs):
        queryset = self.model.objects.all()
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)


class ExcelExportViewSet(viewsets.ModelViewSet):
    model = None

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        queryset = self.get_queryset_by_filter(queryset=self.get_queryset())
        wb = openpyxl.Workbook()
        ws = wb.active

        # # Add headers
        excel_headers = (
            self.model.EXCEL_HEADERS
            if self.model and getattr(self.model, "EXCEL_HEADERS", None)
            else None
        )
        if not excel_headers:
            return Response(
                data={"msg": "Internal Server Error"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        headers = [display_name for _, display_name in excel_headers]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
            ws[f"{col_letter}1"].font = Font(bold=True)

        # # Add data
        for row_num, obj in enumerate(queryset, 2):
            for col_num, (query, header) in enumerate(excel_headers, 1):
                col_letter = get_column_letter(col_num)
                try:
                    value = obj
                    for part in query.split("."):
                        value = getattr(value, part) if value is not None else None
                    if value is None:
                        value = ""
                    elif hasattr(value, "all") and callable(getattr(value, "all")):
                        value = ", ".join(str(item) for item in value.all())
                    ws[f"{col_letter}{row_num}"] = str(value)
                except AttributeError:
                    ws[f"{col_letter}{row_num}"] = ""

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        filename = request.query_params.get("filename")
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        response["File-Name"] = f"{filename}.csv"

        wb.save(response)
        return response

    class Meta:
        abstract = True
